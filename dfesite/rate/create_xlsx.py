import os
# import sys
import openpyxl
from datetime import datetime, timedelta
import dateparser

from django.core.exceptions import ObjectDoesNotExist
from django.conf import settings

MEDIA = settings.MEDIA_DIR
from .models import Daily, Monthly

def sheet_daily(file_path, to_day):
    """ Ежедневно заполняем лист с курсом доллара
    :return: Значение последней даты указанной помесячно
    """
    wb = openpyxl.load_workbook(filename=file_path)
    ws = wb.worksheets[0]
    ws_date = dateparser.parse(ws.cell(row=2, column=1).value, date_formats=['%d.%m.%Y']).date()
    print(f'file_path={file_path}')
    print(f'ws_date={ws_date}')
    print(f'to_day={to_day}')
    ws_monthly = wb.worksheets[1]
    ws_month_date = dateparser.parse(ws_monthly.cell(row=2, column=1).value).date()
    print(f"date__year={ws_date.year}, date__month={ws_date.month}, date__day={ws_date.day}")
    while to_day > ws_date:
        ws_date += timedelta(days=1)
        print('==========WHILE==========')
        print(f"date__year={ws_date.year}, date__month={ws_date.month}, date__day={ws_date.day}")
        try:
            db_daily = Daily.objects.get(date__year=ws_date.year, date__month=ws_date.month, date__day=ws_date.day)
        except ObjectDoesNotExist:
            break
        ws.insert_rows(1, amount=1)
        ws.cell(row=1, column=1).value = 'Дата'
        ws.cell(row=1, column=2).value = 'Курс $'
        ws.cell(row=2, column=1).value = datetime.strptime(str(db_daily.date), "%Y-%m-%d").strftime("%d.%m.%Y")
        ws.cell(row=2, column=2).value = db_daily.usd
        wb.save(filename=file_path)

    return ws_month_date


def sheet_monthly(file_path, db_last):
    wb = openpyxl.load_workbook(filename=file_path)
    ws = wb.worksheets[1]
    ws.insert_rows(1, amount=1)
    ws.cell(row=1, column=1).value = 'Дата'
    ws.cell(row=1, column=2).value = 'Курс $'
    ws.cell(row=1, column=3).value = 'Юралс, $'
    ws.cell(row=2, column=1).value = datetime.strptime(str(db_last.date), "%Y-%m-%d").strftime("%m.%Y")
    ws.cell(row=2, column=2).value = db_last.usd
    ws.cell(row=2, column=3).value = db_last.oil
    wb.save(filename=file_path)


def xl_insert():
    today = datetime.today().date()
    location = os.path.join(MEDIA, 'rate', 'usd_urals.xlsx')
    db_monthly = Monthly.objects.last()
    ws_monthly_date = sheet_daily(location, today)
    ws_date_int = int(str(ws_monthly_date.year) + str(ws_monthly_date.month).zfill(2))
    db_date_int = int(str(db_monthly.date.year) + str(db_monthly.date.month).zfill(2))
    while ws_date_int < db_date_int:
        db_date_as_ws = Monthly.objects.filter(date__year=ws_monthly_date.year,
                                               date__month=ws_monthly_date.month)
        if len(db_date_as_ws):
            sheet_monthly(location, db_date_as_ws[0].get_next_by_date())
        ws_monthly_date = sheet_daily(location, today)
        ws_date_int = int(str(ws_monthly_date.year) + str(ws_monthly_date.month).zfill(2))
    return location
